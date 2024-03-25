import openpyxl
from django.http import HttpResponse


def export_to_xlsx(model_class):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={model_class.__name__}.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = model_class.__name__

    columns = [field.name for field in model_class._meta.fields]

    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title

    rows = model_class.objects.all().values_list(*columns)

    for row_num, row in enumerate(rows, 2):
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value

    wb.save(response)

    return response
